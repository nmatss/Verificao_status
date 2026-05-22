"""Reads the certification spreadsheet and returns all products (any situação)."""

import logging
import os
import warnings
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from .cert_rules import (
    derive_cert_status,
    derive_comercializacao_status,
    derive_license_status,
    stringify_raw,
)
from .config import (
    ENCERRAMENTOS_COLS,
    ENCERRAMENTOS_SHEET,
    EXCEL_FILE,
    SHEET_CONFIG,
)
from .models import Brand, Product

logger = logging.getLogger(__name__)

BRAND_MAP = {
    "Imaginarium": Brand.IMAGINARIUM,
    "Puket": Brand.PUKET,
    "Puket escolares": Brand.PUKET_ESCOLARES,
}

# Headers esperados na linha 1 de cada aba — log warning se divergir.
# Não bloqueia leitura (planilha do fiscal pode ter pequenas variações).
EXPECTED_HEADERS: Dict[str, Dict[int, str]] = {
    "Imaginarium": {
        3: "CÓDIGO",
        6: "NOME",
        21: "SITUAÇÃO",
        22: "Descrição E-commerce",
        23: "Prazo Final Venda",
    },
    "Puket": {
        3: "CÓDIGO",
        6: "NOME",
        20: "SITUAÇÃO",
        21: "Descrição E-commerce",
        22: "Prazo Final Venda",
    },
    "Puket escolares": {
        1: "SKU",
        2: "NOME COMERCIAL",
        7: "STATUS",
        8: "Descrição E-commerce",
        10: "Prazo Final Venda",
    },
}


def _validate_headers(ws, sheet_name: str) -> None:
    """Verifica que os headers da linha 1 batem com o esperado.

    Apenas loga warning quando diverge — não bloqueia.
    """
    expected = EXPECTED_HEADERS.get(sheet_name, {})
    for col, name in expected.items():
        actual = ws.cell(row=1, column=col).value
        actual_str = str(actual or "").strip().lower()
        if actual_str != name.lower():
            logger.warning(
                "Header inesperado em '%s' coluna %d: esperava '%s', achei '%s'.",
                sheet_name, col, name, actual,
            )


# Cache em memória keyed por (path, mtime, brand_filter). Invalida automaticamente
# quando o arquivo é editado (mtime muda).
_CACHE: Dict[Tuple[str, float, Any], List[Product]] = {}


def _read_cell(ws, row_idx: int, col_idx: Optional[int]) -> Any:
    """Lê o valor cru de uma célula respeitando colunas opcionais (None)."""
    if col_idx is None:
        return None
    return ws.cell(row=row_idx, column=col_idx).value


def _coerce_estoque(value: Any) -> Optional[int]:
    """Converte estoque informado em int. Retorna None se vazio/inválido."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s.replace(",", ".")))
        except (ValueError, OverflowError):
            return None
    return None


def _coerce_ean(value: Any) -> Optional[str]:
    """Converte EAN em string limpa (sem .0 trailing). None se vazio."""
    if value is None:
        return None
    if isinstance(value, float):
        # Excel devolve EAN como float; preserva como inteiro string
        if value != value:  # NaN
            return None
        try:
            return str(int(value))
        except (ValueError, OverflowError):
            return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return None


def _read_encerramentos(wb) -> Dict[str, Dict[str, Any]]:
    """Lê a aba Encerramentos e retorna {sku: {codigo_barras, estoque_informado, prazo_final_venda_raw}}.

    A coluna G (prazo_final_venda) é mantida pela fiscal e tem prioridade sobre
    o prazo da aba da marca quando o produto está em encerramento.

    Se a aba não existir, loga warning e retorna dict vazio.
    """
    if ENCERRAMENTOS_SHEET not in wb.sheetnames:
        logger.warning(
            "Aba '%s' não encontrada na planilha — produtos ficarão sem "
            "EAN/estoque enriquecido.",
            ENCERRAMENTOS_SHEET,
        )
        return {}

    ws = wb[ENCERRAMENTOS_SHEET]
    out: Dict[str, Dict[str, Any]] = {}

    prazo_col = ENCERRAMENTOS_COLS.get("prazo_final_venda")

    for row_idx in range(2, ws.max_row + 1):
        sku_val = ws.cell(row=row_idx, column=ENCERRAMENTOS_COLS["sku"]).value
        if not sku_val:
            continue
        sku = str(sku_val).strip()
        if not sku:
            continue

        ean_raw = ws.cell(row=row_idx, column=ENCERRAMENTOS_COLS["codigo_barras"]).value
        estoque_raw = ws.cell(
            row=row_idx, column=ENCERRAMENTOS_COLS["estoque_informado"]
        ).value
        prazo_raw = ws.cell(row=row_idx, column=prazo_col).value if prazo_col else None

        out[sku] = {
            "codigo_barras": _coerce_ean(ean_raw),
            "estoque_informado": _coerce_estoque(estoque_raw),
            "prazo_final_venda_raw": prazo_raw,
        }

    return out


def read_products(excel_path=None, brand_filter=None) -> List[Product]:
    """Read all products from the Excel spreadsheet (independent of situação).

    Args:
        excel_path: Path to Excel file. Defaults to configured path.
        brand_filter: Optional Brand enum to filter by.

    Returns:
        List of Product instances (all situações: Ativo, Encerrado, etc.).
    """
    path_str = str(excel_path or EXCEL_FILE)

    try:
        mtime = os.path.getmtime(path_str)
    except OSError:
        mtime = 0.0

    cache_key = (path_str, mtime, brand_filter)
    cached = _CACHE.get(cache_key)
    if cached is not None:
        return cached

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(path_str, data_only=True)

    # Lê o mapa de Encerramentos uma vez só (independente de brand_filter)
    encerramentos_map = _read_encerramentos(wb)

    products: List[Product] = []

    for sheet_name, cols in SHEET_CONFIG.items():
        brand = BRAND_MAP[sheet_name]

        if brand_filter and brand != brand_filter:
            # If filtering by "puket", include both Puket and Puket escolares
            if brand_filter == Brand.PUKET and brand == Brand.PUKET_ESCOLARES:
                pass
            elif brand_filter == Brand.PUKET_ESCOLARES and brand == Brand.PUKET:
                pass
            else:
                continue

        ws = wb[sheet_name]
        _validate_headers(ws, sheet_name)

        for row_idx in range(2, ws.max_row + 1):
            sku = ws.cell(row=row_idx, column=cols["sku"]).value
            if not sku:
                continue  # Skip header/family rows without SKU

            sku = str(sku).strip()
            if not sku:
                continue

            name = ws.cell(row=row_idx, column=cols["name"]).value
            name = str(name).strip() if name else ""

            cert_text = ws.cell(row=row_idx, column=cols["cert_text"]).value
            cert_text = str(cert_text).strip() if cert_text else None

            situacao_raw = _read_cell(ws, row_idx, cols.get("status"))
            situacao = stringify_raw(situacao_raw)

            tipo_raw = _read_cell(ws, row_idx, cols.get("tipo_cert"))
            tipo_cert = stringify_raw(tipo_raw)

            validade_raw = _read_cell(ws, row_idx, cols.get("validade_cert"))
            validade_str = stringify_raw(validade_raw)

            prazo_raw = _read_cell(ws, row_idx, cols.get("prazo_final_venda"))
            prazo_str_marca = stringify_raw(prazo_raw)

            numero_raw = _read_cell(ws, row_idx, cols.get("numero_registro"))
            numero_registro = stringify_raw(numero_raw)

            # Merge enriquecimento da aba Encerramentos (por SKU)
            extra = encerramentos_map.get(sku, {})
            codigo_barras = extra.get("codigo_barras")
            estoque_informado = extra.get("estoque_informado")
            prazo_encerramentos_raw = extra.get("prazo_final_venda_raw")
            prazo_encerramentos_str = stringify_raw(prazo_encerramentos_raw)

            # Regra solicitada pela Carla:
            #   - Se Encerramentos tem prazo → usa (prioridade da fiscal)
            #   - Senão, se aba da marca tem → usa
            #   - Senão None
            if prazo_encerramentos_str:
                prazo_effective_raw = prazo_encerramentos_raw
                prazo_str_final = prazo_encerramentos_str
            else:
                prazo_effective_raw = prazo_raw
                prazo_str_final = prazo_str_marca

            # Deriva DEPOIS do merge — prazo influencia cert_status/comercializacao
            cert_status = derive_cert_status(situacao, prazo_effective_raw)
            license_status = derive_license_status(tipo_cert, validade_raw)
            comercializacao_status = derive_comercializacao_status(
                cert_status, situacao, prazo_effective_raw
            )

            products.append(Product(
                sku=sku,
                name=name,
                brand=brand,
                expected_cert_text=cert_text,
                excel_row=row_idx,
                situacao=situacao,
                tipo_certificacao=tipo_cert,
                validade_certificacao_raw=validade_str,
                prazo_final_venda_raw=prazo_str_final,
                numero_registro=numero_registro,
                codigo_barras=codigo_barras,
                estoque_informado=estoque_informado,
                cert_status=cert_status,
                license_status=license_status,
                comercializacao_status=comercializacao_status,
            ))

    wb.close()
    _CACHE[cache_key] = products
    return products
