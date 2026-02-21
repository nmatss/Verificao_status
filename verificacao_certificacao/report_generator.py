"""Report generation in Excel and CSV formats."""

import csv
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .config import REPORTS_DIR
from .models import ValidationResult, ValidationStatus

# Color fills for conditional formatting
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_MISSING = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_INCONSISTENT = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ERROR = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(bold=True, size=11)

STATUS_FILLS = {
    ValidationStatus.OK: FILL_OK,
    ValidationStatus.MISSING: FILL_MISSING,
    ValidationStatus.INCONSISTENT: FILL_INCONSISTENT,
    ValidationStatus.URL_NOT_FOUND: FILL_ERROR,
    ValidationStatus.API_ERROR: FILL_ERROR,
    ValidationStatus.NO_EXPECTED: FILL_ERROR,
}

COLUMNS = [
    "SKU", "Nome", "Marca", "Aba", "Texto Esperado",
    "Texto Real", "Status", "Score", "URL", "Erro",
    "Avaliação IA",
]


def generate_reports(results: List[ValidationResult], output_dir: Path = None) -> Path:
    """Generate Excel and CSV reports.

    Args:
        results: List of validation results.
        output_dir: Output directory. Defaults to reports/.

    Returns:
        Path to the generated Excel file.
    """
    output_dir = output_dir or REPORTS_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = output_dir / f"validacao_{timestamp}.xlsx"
    csv_path = output_dir / f"validacao_{timestamp}.csv"

    _generate_excel(results, excel_path)
    _generate_csv(results, csv_path)

    return excel_path


def _generate_excel(results: List[ValidationResult], path: Path):
    """Generate Excel report with conditional formatting."""
    wb = openpyxl.Workbook()

    # === Detail sheet ===
    ws = wb.active
    ws.title = "Resultados"

    # Headers
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, result in enumerate(results, 2):
        p = result.product
        row_data = [
            p.sku,
            p.name,
            p.brand.value,
            p.brand.value,
            p.expected_cert_text or "",
            result.actual_cert_text or "",
            result.status.value,
            round(result.similarity_score, 2),
            p.resolved_url or "",
            result.error_message or "",
            result.ai_assessment or "",
        ]

        fill = STATUS_FILLS.get(result.status, FILL_ERROR)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Apply fill to the status column and the whole row
            cell.fill = fill

    # Auto-adjust column widths
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(COLUMNS[col_idx - 1])
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    # === Summary sheet ===
    ws_summary = wb.create_sheet("Resumo")
    _write_summary(ws_summary, results)

    wb.save(str(path))


def _write_summary(ws, results: List[ValidationResult]):
    """Write summary tab with counts by brand and status."""
    ws.cell(row=1, column=1, value="Resumo da Validação").font = FONT_BOLD

    # Overall counts
    ws.cell(row=3, column=1, value="Status").font = FONT_BOLD
    ws.cell(row=3, column=2, value="Quantidade").font = FONT_BOLD
    ws.cell(row=3, column=3, value="Percentual").font = FONT_BOLD

    total = len(results)
    status_counts = {}
    for status in ValidationStatus:
        count = sum(1 for r in results if r.status == status)
        if count > 0:
            status_counts[status] = count

    row = 4
    for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=status.value)
        ws.cell(row=row, column=1).fill = STATUS_FILLS.get(status, FILL_ERROR)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = FONT_BOLD
    ws.cell(row=row, column=2, value=total).font = FONT_BOLD

    # Per-brand breakdown
    row += 2
    ws.cell(row=row, column=1, value="Por Marca").font = FONT_BOLD
    row += 1

    brands = sorted(set(r.product.brand.value for r in results))
    headers = ["Marca", "Total"] + [s.value for s in ValidationStatus]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    row += 1
    for brand_name in brands:
        brand_results = [r for r in results if r.product.brand.value == brand_name]
        ws.cell(row=row, column=1, value=brand_name)
        ws.cell(row=row, column=2, value=len(brand_results))
        for status_idx, status in enumerate(ValidationStatus, 3):
            count = sum(1 for r in brand_results if r.status == status)
            ws.cell(row=row, column=status_idx, value=count)
        row += 1

    # Auto-adjust widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18


def _generate_csv(results: List[ValidationResult], path: Path):
    """Generate CSV report."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(COLUMNS)

        for result in results:
            p = result.product
            writer.writerow([
                p.sku,
                p.name,
                p.brand.value,
                p.brand.value,
                p.expected_cert_text or "",
                result.actual_cert_text or "",
                result.status.value,
                round(result.similarity_score, 2),
                p.resolved_url or "",
                result.error_message or "",
                result.ai_assessment or "",
            ])
